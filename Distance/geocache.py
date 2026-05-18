"""
geocache.py — Geocoding cache using Excel file

Logic:
    - On first run: geocode all addresses, save to geocache.xlsx
    - On next runs: read from geocache.xlsx, only geocode NEW addresses
    - This saves time and API calls

geocache.xlsx structure:
    | name | address | lat | lon |
"""

import requests
import os

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.run(["pip", "install", "openpyxl", "--break-system-packages", "-q"])
    import openpyxl

# ── Config ────────────────────────────────────────────────────────────────────
ORS_API_KEY  = "eyJvcmciOiI1YjNjZTM1OTc4NTExMTAwMDFjZjYyNDgiLCJpZCI6Ijc4MzMyNTBkYTBhMDRiYTg5MDQyYzkxNTQ4MDY0MzQ4IiwiaCI6Im11cm11cjY0In0="
CACHE_FILE   = "geocache.xlsx"


# ── Load cache from Excel ─────────────────────────────────────────────────────

def load_cache() -> dict:
    """
    Read geocache.xlsx and return a dict:
        { address_string: {"lat": float, "lon": float} }
    """
    cache = {}

    if not os.path.exists(CACHE_FILE):
        return cache  # empty cache, file doesn't exist yet

    wb  = openpyxl.load_workbook(CACHE_FILE)
    ws  = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
        name, address, lat, lon = row
        if address and lat and lon:
            cache[address] = {"name": name, "lat": float(lat), "lon": float(lon)}

    print(f"  📂 Loaded {len(cache)} cached addresses from {CACHE_FILE}")
    return cache


# ── Save cache to Excel ───────────────────────────────────────────────────────

def save_cache(cache: dict):
    """
    Write the full cache dict back to geocache.xlsx.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Geocache"

    # Header
    ws.append(["name", "address", "lat", "lon"])

    # Data rows
    for address, data in cache.items():
        ws.append([data["name"], address, data["lat"], data["lon"]])

    # Column widths for readability
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15

    wb.save(CACHE_FILE)


# ── Geocode one address (API call) ────────────────────────────────────────────

def geocode_api(name: str, address: str) -> dict:
    """
    Call ORS Geocoding API to convert address → lat/lon.
    Only called when address is NOT in cache.
    """
    url     = "https://api.openrouteservice.org/geocode/search"
    headers = {"Authorization": ORS_API_KEY}
    params  = {"text": address, "boundary.country": "FR", "size": 1}

    response = requests.get(url, headers=headers, params=params)

    if response.status_code != 200 or not response.json()["features"]:
        print(f"  ❌ Could not geocode: {address}")
        return None

    feature  = response.json()["features"][0]
    lon, lat = feature["geometry"]["coordinates"]

    return {"name": name, "lat": lat, "lon": lon}


# ── Main geocode function (cache-aware) ───────────────────────────────────────

def geocode(name: str, address: str, cache: dict) -> dict:
    """
    Geocode an address using cache first, API only if not found.

    Returns: {"name": str, "address": str, "lat": float, "lon": float}
    """
    # Check cache first
    if address in cache:
        data = cache[address]
        print(f"  ✅ Cache hit:  {name} → ({data['lat']:.4f}, {data['lon']:.4f})")
        return {"name": name, "address": address, "lat": data["lat"], "lon": data["lon"]}

    # Not in cache → call API
    print(f"  🌐 Geocoding:  {name} → {address}")
    result = geocode_api(name, address)

    if result:
        # Save to cache for next time
        cache[address] = {"name": name, "lat": result["lat"], "lon": result["lon"]}
        return {"name": name, "address": address, "lat": result["lat"], "lon": result["lon"]}

    return None


def geocode_all(orders: list, warehouse: dict) -> list:
    """
    Geocode warehouse + all orders using cache.
    Saves any new addresses back to cache file automatically.

    Returns: list of location dicts with lat/lon added
    """
    print("\n" + "="*65)
    print("  GEOCODING (with cache)")
    print("="*65)

    # Load existing cache
    cache = load_cache()
    new_count = 0

    # Geocode warehouse
    wh = geocode(warehouse["name"], warehouse["address"], cache)
    if wh is None:
        print("  ERROR: Could not geocode warehouse.")
        exit(1)
    wh["id"]          = "W0"
    wh["priority"]    = None
    wh["time_window"] = None
    wh["boxes"]       = 0
    wh["is_cold"]     = False
    wh["is_suburban"] = False

    # Geocode all orders
    locations = [wh]
    for order in orders:
        loc = geocode(order["name"], order["address"], cache)
        if loc:
            loc["id"]          = order["id"]
            loc["priority"]    = order["priority"]
            loc["time_window"] = order["time_window"]
            loc["boxes"]       = order["boxes"]
            loc["is_cold"]     = order["is_cold"]
            loc["is_suburban"] = order["is_suburban"]
            locations.append(loc)

    # Count new addresses and save if any
    new_count = len(cache) - (len(locations) - 1)  # rough estimate
    save_cache(cache)
    print(f"\n  💾 Cache saved to {CACHE_FILE} ({len(cache)} total addresses)")

    return locations


# ── Test ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    # Quick test with a few addresses
    test_orders = [
        {"id": "T1", "name": "Hospital Saint-Antoine", "address": "184 Rue du Faubourg Saint-Antoine, Paris",
         "priority": "critical", "is_cold": False, "is_suburban": False, "boxes": 5, "time_window": ("06:30", "09:00")},
        {"id": "T2", "name": "Pharmacie Bastille",     "address": "6 Place de la Bastille, Paris",
         "priority": "high",     "is_cold": False, "is_suburban": False, "boxes": 2, "time_window": ("08:00", "11:00")},
    ]
    warehouse = {"id": "W0", "name": "Warehouse", "address": "14 Rue du Chemin Vert, Paris"}

    print("=== FIRST RUN (should geocode via API) ===")
    locations = geocode_all(test_orders, warehouse)

    print("\n=== SECOND RUN (should use cache) ===")
    locations = geocode_all(test_orders, warehouse)

    print(f"\nDone. Check {CACHE_FILE} to see saved coordinates.")
