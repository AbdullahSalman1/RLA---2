"""
routing.py — Route sequencing using Clarke-Wright Savings + 2-opt improvement
Built on top of assign.py

Priority Rule (from client interview):
    Priority beats proximity. Always. No exceptions.
    Critical stops → first (sorted by earliest deadline)
    High stops     → second (sorted by earliest deadline)
    Normal stops   → Clarke-Wright + 2-opt (distance optimization only)

Pipeline:
    1. Geocode all addresses
    2. Get real road distance matrix (ORS API)
    3. Assign orders to vehicles
    4. For each vehicle:
        a. Lock critical stops at front (sorted by deadline)
        b. Lock high stops after critical (sorted by deadline)
        c. Run Clarke-Wright + 2-opt on normal stops only
        d. Combine: critical + high + optimized normal
    5. Simulate route timing, flag late deliveries
    6. Print final plan

Run:
    python routing.py
"""

from pathlib import Path

import requests
from openpyxl import load_workbook
from geocache import geocode_all as geocode_all_cached

# ── API Key ───────────────────────────────────────────────────────────────────
ORS_API_KEY = "eyJvcmciOiI1YjNjZTM1OTc4NTExMTAwMDFjZjYyNDgiLCJpZCI6Ijc4MzMyNTBkYTBhMDRiYTg5MDQyYzkxNTQ4MDY0MzQ4IiwiaCI6Im11cm11cjY0In0="

# ── Fleet ─────────────────────────────────────────────────────────────────────
VEHICLES = [
    {"id": "V1", "type": "small",        "label": "Small Van 1",      "max_stops": 30},
    {"id": "V2", "type": "small",        "label": "Small Van 2",      "max_stops": 30},
    {"id": "V3", "type": "small",        "label": "Small Van 3",      "max_stops": 30},
    {"id": "V4", "type": "large",        "label": "Large Van 1",      "max_stops": 20},
    {"id": "V5", "type": "large",        "label": "Large Van 2",      "max_stops": 20},
    {"id": "V6", "type": "refrigerated", "label": "Refrigerated Van", "max_stops": 15},
]

# ── Drivers ───────────────────────────────────────────────────────────────────
DRIVERS = [
    {"id": "D1", "name": "Alexandre M.", "certified_refrigerated": True,  "shift_start": "06:30", "max_hours": 8},
    {"id": "D2", "name": "Fatima B.",    "certified_refrigerated": True,  "shift_start": "06:30", "max_hours": 8},
    {"id": "D3", "name": "Thomas L.",    "certified_refrigerated": False, "shift_start": "07:00", "max_hours": 8},
    {"id": "D4", "name": "Yasmine K.",   "certified_refrigerated": False, "shift_start": "07:00", "max_hours": 8},
    {"id": "D5", "name": "Romain D.",    "certified_refrigerated": False, "shift_start": "07:00", "max_hours": 8},
]

DATA_FILE = Path(__file__).with_name("delivery_data.xlsx")


def _as_bool(value):
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "y", "oui"}


def load_delivery_data(path=DATA_FILE):
    """Load warehouse and orders from delivery_data.xlsx."""
    if not path.exists():
        raise FileNotFoundError(
            f"Missing {path.name}. Create it with warehouse and orders sheets."
        )

    wb = load_workbook(path)

    if "warehouse" not in wb.sheetnames or "orders" not in wb.sheetnames:
        raise ValueError(
            f"{path.name} must contain 'warehouse' and 'orders' sheets."
        )

    ws_wh = wb["warehouse"]
    wh_rows = list(ws_wh.iter_rows(values_only=True))
    wh_headers = [str(v).strip().lower() for v in wh_rows[0]]
    wh_data = wh_rows[1]
    wh_map = {name: idx for idx, name in enumerate(wh_headers)}

    warehouse = {
        "id": str(wh_data[wh_map["id"]]),
        "name": str(wh_data[wh_map["name"]]),
        "address": str(wh_data[wh_map["address"]]),
    }

    ws_orders = wb["orders"]
    order_rows = list(ws_orders.iter_rows(values_only=True))
    order_headers = [str(v).strip().lower() for v in order_rows[0]]
    order_map = {name: idx for idx, name in enumerate(order_headers)}

    orders = []
    for row in order_rows[1:]:
        if not any(row):
            continue
        orders.append({
            "id": str(row[order_map["id"]]),
            "name": str(row[order_map["name"]]),
            "address": str(row[order_map["address"]]),
            "priority": str(row[order_map["priority"]]),
            "is_cold": _as_bool(row[order_map["is_cold"]]),
            "is_suburban": _as_bool(row[order_map["is_suburban"]]),
            "boxes": int(row[order_map["boxes"]]),
            "time_window": (
                str(row[order_map["time_window_start"]]),
                str(row[order_map["time_window_end"]]),
            ),
        })

    return warehouse, orders


WAREHOUSE, ORDERS = load_delivery_data()

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 1 — GEOCODING
# ═══════════════════════════════════════════════════════════════════════════════

def geocode(name, address):
    """Convert address to lat/lon using ORS Geocoding API."""
    url     = "https://api.openrouteservice.org/geocode/search"
    headers = {"Authorization": ORS_API_KEY}
    params  = {"text": address, "boundary.country": "FR", "size": 1}

    response = requests.get(url, headers=headers, params=params)
    if response.status_code != 200 or not response.json()["features"]:
        print(f"  Could not geocode: {address}")
        return None

    feature  = response.json()["features"][0]
    lon, lat = feature["geometry"]["coordinates"]
    return {"name": name, "address": address, "lat": lat, "lon": lon}


def geocode_all(orders):
    """Geocode warehouse and all orders."""
    print("\n" + "="*65)
    print("  STEP 1 — GEOCODING ADDRESSES")
    print("="*65)

    wh = geocode(WAREHOUSE["name"], WAREHOUSE["address"])
    wh["id"]          = "W0"
    wh["priority"]    = None
    wh["time_window"] = None
    wh["boxes"]       = 0
    wh["is_cold"]     = False
    wh["is_suburban"] = False
    print(f"  ✓ Warehouse: ({wh['lat']:.4f}, {wh['lon']:.4f})")

    locations = [wh]
    for order in orders:
        loc = geocode(order["name"], order["address"])
        if loc:
            loc["id"]          = order["id"]
            loc["priority"]    = order["priority"]
            loc["time_window"] = order["time_window"]
            loc["boxes"]       = order["boxes"]
            loc["is_cold"]     = order["is_cold"]
            loc["is_suburban"] = order["is_suburban"]
            locations.append(loc)
            print(f"  ✓ {order['name']}: ({loc['lat']:.4f}, {loc['lon']:.4f})")

    return locations


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 2 — DISTANCE MATRIX
# ═══════════════════════════════════════════════════════════════════════════════

def get_distance_matrix(locations):
    """Get real road distances between all locations in one API call."""
    print("\n" + "="*65)
    print("  STEP 2 — FETCHING DISTANCE MATRIX")
    print("="*65)

    url     = "https://api.openrouteservice.org/v2/matrix/driving-car"
    headers = {"Authorization": ORS_API_KEY, "Content-Type": "application/json"}
    coords  = [[loc["lon"], loc["lat"]] for loc in locations]

    body = {
        "locations": coords,
        "metrics":   ["distance", "duration"],
        "units":     "km"
    }

    print(f"  Calling ORS Matrix API with {len(locations)} locations...")
    response = requests.post(url, headers=headers, json=body)

    if response.status_code != 200:
        print(f"  API Error {response.status_code}: {response.text}")
        return None, None

    data      = response.json()
    distances = data["distances"]
    durations = [[s / 60 for s in row] for row in data["durations"]]

    print(f"  ✓ Matrix ready: {len(locations)}x{len(locations)} pairs")
    return distances, durations


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 3 — VEHICLE ASSIGNMENT
# ═══════════════════════════════════════════════════════════════════════════════

def assign_orders_to_vehicles(orders):
    """
    Assign each order to the correct vehicle type.
    Cold chain  → refrigerated van (hard constraint)
    Suburban    → large van (range)
    City        → small van (agile)
    """
    priority_rank  = {"critical": 0, "high": 1, "normal": 2}
    sorted_orders  = sorted(orders, key=lambda o: priority_rank[o["priority"]])
    vehicle_stops  = {v["id"]: 0 for v in VEHICLES}
    vehicle_orders = {v["id"]: [] for v in VEHICLES}

    print("\n" + "="*65)
    print("  STEP 3 — ASSIGNING ORDERS TO VEHICLES")
    print("="*65)

    for order in sorted_orders:
        vid    = None
        reason = ""

        if order["is_cold"]:
            for v in VEHICLES:
                if v["type"] == "refrigerated" and vehicle_stops[v["id"]] < v["max_stops"]:
                    vid    = v["id"]
                    reason = "cold chain → refrigerated"
                    break

        elif order["is_suburban"]:
            for v in VEHICLES:
                if v["type"] == "large" and vehicle_stops[v["id"]] < v["max_stops"]:
                    vid    = v["id"]
                    reason = "suburban → large van"
                    break
            if not vid:
                for v in VEHICLES:
                    if v["type"] == "small" and vehicle_stops[v["id"]] < v["max_stops"]:
                        vid    = v["id"]
                        reason = "suburban (large full) → small van"
                        break
        else:
            for v in VEHICLES:
                if v["type"] == "small" and vehicle_stops[v["id"]] < v["max_stops"]:
                    vid    = v["id"]
                    reason = "city → small van"
                    break

        if vid:
            vehicle_orders[vid].append(order)
            vehicle_stops[vid] += 1
            vlabel = next(v["label"] for v in VEHICLES if v["id"] == vid)
            print(f"  {order['id']} {order['name']:<28} → {vlabel} ({reason})")
        else:
            print(f"  {order['id']} {order['name']:<28} → ⚠ UNASSIGNED")

    return vehicle_orders


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 4 — DRIVER ASSIGNMENT
# ═══════════════════════════════════════════════════════════════════════════════

def assign_drivers(vehicle_orders):
    """
    Refrigerated van → certified driver only.
    All other vans   → any available driver.
    """
    assigned_drivers  = set()
    driver_assignment = {}

    # First pass: certified driver to refrigerated van
    for v in VEHICLES:
        if v["type"] == "refrigerated" and vehicle_orders.get(v["id"]):
            for d in DRIVERS:
                if d["certified_refrigerated"] and d["id"] not in assigned_drivers:
                    driver_assignment[v["id"]] = d
                    assigned_drivers.add(d["id"])
                    break

    # Second pass: remaining drivers to other vehicles
    remaining = [d for d in DRIVERS if d["id"] not in assigned_drivers]
    idx = 0
    for v in VEHICLES:
        if vehicle_orders.get(v["id"]) and v["id"] not in driver_assignment:
            if idx < len(remaining):
                driver_assignment[v["id"]] = remaining[idx]
                idx += 1

    return driver_assignment


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 5 — ROUTING: PRIORITY LOCKING + CLARKE-WRIGHT + 2-OPT
# ═══════════════════════════════════════════════════════════════════════════════

def time_to_min(t):
    h, m = map(int, t.split(":"))
    return h * 60 + m

def min_to_time(m):
    return f"{int(m) // 60:02d}:{int(m) % 60:02d}"


def simulate_arrival_times(route, locations, warehouse_idx=0):
    """
    Simulate arrival times for a given route order.
    Uses average city speed of 25 km/h as estimate.
    Returns dict: { stop_index: arrival_minutes_since_midnight }

    Note: this is a lightweight estimate used only for time window checking,
    not the full simulation (which uses real durations from ORS matrix).
    We use 07:00 as default start since normal stops are on non-fridge,
    non-critical vehicles whose drivers start at 07:00.
    """
    AVERAGE_SPEED_KMH = 25
    SERVICE_MIN       = 10
    current_min       = 7 * 60   # 07:00 default start
    prev_idx          = warehouse_idx
    arrivals          = {}

    for idx in route:
        # Estimate travel time from previous stop
        # We use a placeholder distance here — real durations used in final sim
        # This is just to check ORDER correctness, not exact times
        travel_min      = 15   # conservative 15-min estimate per stop
        arrival_min     = current_min + travel_min
        arrivals[idx]   = arrival_min
        current_min     = arrival_min + SERVICE_MIN
        prev_idx        = idx

    return arrivals


def fix_time_windows(route, locations):
    """
    After Clarke-Wright + 2-opt, only move a stop earlier if it WILL be late
    in the current order.

    Rule:
        If driver CAN reach C on time in current order → keep A → B → C
        If driver CANNOT reach C on time              → move C to earliest
                                                         position where it fits

    This preserves distance optimization whenever possible.
    Only intervenes when a real deadline violation is detected.
    """
    if not route:
        return route

    improved = True

    while improved:
        improved = False

        # Simulate arrival times in current order
        arrivals = simulate_arrival_times(route, locations)

        for i, idx in enumerate(route):
            tw_end      = time_to_min(locations[idx]["time_window"][1])
            arrival_min = arrivals[idx]

            # Is this stop going to be late in current position?
            if arrival_min > tw_end:
                # Try moving it to every earlier position
                for j in range(i):
                    new_route    = route[:j] + [idx] + route[j:i] + route[i+1:]
                    new_arrivals = simulate_arrival_times(new_route, locations)

                    # Does moving to position j fix the lateness?
                    if new_arrivals[idx] <= tw_end:
                        print(f"    ⏰ Time window fix: moved {locations[idx]['name']} "
                              f"from position {i+1} to {j+1} "
                              f"(closes {locations[idx]['time_window'][1]}, "
                              f"was arriving at {min_to_time(int(arrival_min))})")
                        route    = new_route
                        improved = True
                        break

                if improved:
                    break  # restart the scan with updated route

    return route


def build_route(stop_indices, locations, distances):
    """
    Build the delivery sequence for one vehicle.

    Rule: Priority beats proximity. Always.
        Step 1 — Lock critical stops at front, sorted by deadline
        Step 2 — Lock high stops after critical, sorted by deadline
        Step 3 — Run Clarke-Wright + 2-opt on normal stops only
        Step 4 — Return: critical + high + optimized_normal
    """

    # Separate stops by priority
    critical = [i for i in stop_indices if locations[i]["priority"] == "critical"]
    high     = [i for i in stop_indices if locations[i]["priority"] == "high"]
    normal   = [i for i in stop_indices if locations[i]["priority"] == "normal"]

    # Sort critical and high by earliest deadline — tightest window first
    critical.sort(key=lambda i: (time_to_min(locations[i]["time_window"][0]), time_to_min(locations[i]["time_window"][1])))
    high.sort(key=lambda i: (time_to_min(locations[i]["time_window"][0]), time_to_min(locations[i]["time_window"][1])))

    # Run Clarke-Wright only on normal stops
    normal_routed = clarke_wright(normal, distances, warehouse_idx=0)

    # Run 2-opt only on normal stops — never touch locked stops
    normal_routed = two_opt(normal_routed, distances)

    # Fix time windows: if Clarke-Wright put C last but C closes before A,
    # move C to the earliest position where it won't be late
    normal_routed = fix_time_windows(normal_routed, locations)

    # Final route: locked critical → locked high → optimized normal
    final_route = critical + high + normal_routed

    return final_route, len(critical) + len(high)


def clarke_wright(stop_indices, distances, warehouse_idx=0):
    """
    Clarke-Wright Savings Algorithm.

    saving(i,j) = dist(W→i) + dist(W→j) - dist(i→j)
    Higher saving = more benefit to visiting i and j in the same trip.
    Sort savings descending, merge pairs greedily.
    """
    if not stop_indices:
        return []
    if len(stop_indices) == 1:
        return list(stop_indices)

    # Compute savings for every pair
    savings = []
    for i in stop_indices:
        for j in stop_indices:
            if i >= j:
                continue
            saving = (distances[warehouse_idx][i] +
                      distances[warehouse_idx][j] -
                      distances[i][j])
            savings.append((saving, i, j))

    # Sort descending by saving
    savings.sort(reverse=True, key=lambda x: x[0])

    # Merge pairs greedily
    visited = set()
    ordered = []

    for saving, i, j in savings:
        if i not in visited and j not in visited:
            ordered.append(i)
            ordered.append(j)
            visited.add(i)
            visited.add(j)

    # Add any remaining stops not yet paired
    for idx in stop_indices:
        if idx not in visited:
            ordered.append(idx)

    return ordered


def two_opt(route, distances):
    """
    2-opt local search.

    Try reversing every segment between two edges.
    Keep reversal if it reduces total distance.
    Repeat until no improvement found.

    Only runs on the stops passed in — never touches locked priority stops.
    """
    if len(route) <= 2:
        return route

    best     = route[:]
    improved = True

    while improved:
        improved = False
        for i in range(len(best) - 1):
            for j in range(i + 2, len(best)):
                # Distance of current edges
                current = (distances[best[i]][best[i + 1]] +
                           distances[best[j]][best[(j + 1) % len(best)]])

                # Distance after reversing segment between i+1 and j
                new = (distances[best[i]][best[j]] +
                       distances[best[i + 1]][best[(j + 1) % len(best)]])

                if new < current - 0.001:
                    best[i + 1:j + 1] = best[i + 1:j + 1][::-1]
                    improved = True

    return best


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 6 — TIME SIMULATION
# ═══════════════════════════════════════════════════════════════════════════════

SERVICE_TIME_MIN = 10  # average unloading time per stop (minutes)

def simulate_route(route_indices, locations, durations, driver, distances=None):
    """
    Walk through the sequenced route and compute:
    - Arrival time at each stop
    - Whether the delivery is within its time window
    - Whether the driver shift is exceeded
    """
    results     = []
    current_min = time_to_min(driver["shift_start"])
    max_end_min = current_min + driver["max_hours"] * 60
    prev_idx    = 0  # start from warehouse (index 0)

    for idx in route_indices:
        loc        = locations[idx]
        travel_min = durations[prev_idx][idx]
        arrival_min = current_min + travel_min

        # Wait if arriving before time window opens
        tw_start = time_to_min(loc["time_window"][0])
        tw_end   = time_to_min(loc["time_window"][1])

        if arrival_min < tw_start:
            arrival_min = tw_start  # driver waits at location

        is_late       = arrival_min > tw_end
        departure_min = arrival_min + SERVICE_TIME_MIN

        dist_from_prev = distances[prev_idx][idx] if distances else 0.0

        results.append({
            "id":             loc["id"],
            "name":           loc["name"],
            "priority":       loc["priority"],
            "arrival":        min_to_time(arrival_min),
            "departure":      min_to_time(departure_min),
            "window":         f"{loc['time_window'][0]}–{loc['time_window'][1]}",
            "is_late":        is_late,
            "late_by_min":    max(0, int(arrival_min - tw_end)),
            "shift_exceeded": departure_min > max_end_min,
            "dist_from_prev": round(dist_from_prev, 2),
        })

        current_min = departure_min
        prev_idx    = idx

    total_duration = current_min - time_to_min(driver["shift_start"])
    return results, total_duration


# ═══════════════════════════════════════════════════════════════════════════════
# PRINT FINAL PLAN
# ═══════════════════════════════════════════════════════════════════════════════

def print_routes(vehicle_routes):
    print("\n" + "="*70)
    print("  FINAL DELIVERY PLAN")
    print("="*70)

    total_late = 0

    for entry in vehicle_routes:
        v      = entry["vehicle"]
        driver = entry["driver"]
        stops  = entry["stops"]
        km     = entry["total_km"]
        dur    = entry["total_duration_min"]

        if not stops:
            continue

        dname = driver["name"] if driver else "⚠ No driver"
        print(f"\n  🚐 {v['label']}  |  Driver: {dname}  |  {len(stops)} stops  |  {km:.1f} km  |  ~{int(dur)} min")
        print(f"  {'#':<4} {'Client':<28} {'Priority':<10} {'Window':<14} {'Arrival':<10} {'Dist':>8}  {'Status'}")
        print("  " + "-"*80)

        for i, stop in enumerate(stops):
            priority_icon = "🔴" if stop["priority"] == "critical" else ("🟡" if stop["priority"] == "high" else "⚪")
            status        = "🔴 LATE" if stop["is_late"] else "✅ OK"
            shift_warn    = " ⚠ SHIFT EXCEEDED" if stop["shift_exceeded"] else ""
            late_note     = f" (+{stop['late_by_min']} min)" if stop["is_late"] else ""

            dist_str = f"{stop['dist_from_prev']:.1f} km" if i > 0 else "start"
            print(f"  {i+1:<4} {stop['name']:<28} {priority_icon} {stop['priority']:<8} {stop['window']:<14} {stop['arrival']:<10} {dist_str:>8}  {status}{late_note}{shift_warn}")

            if stop["is_late"]:
                total_late += 1

        # Return to warehouse with distance
        if entry.get("return_km") is not None:
            print(f"  {'':<4} {'↩ Return to Warehouse':<28} {'':<10} {'':<14} {'':<10} {entry['return_km']:.1f} km")
        else:
            print(f"  {'':<4} {'↩ Return to Warehouse':<28}")

    print("\n" + "="*70)
    print(f"  SUMMARY")
    print("="*70)
    total_stops = sum(len(e["stops"]) for e in vehicle_routes)
    total_km    = sum(e["total_km"] for e in vehicle_routes)
    print(f"  Total deliveries : {total_stops}")
    print(f"  Total distance   : {total_km:.1f} km")
    print(f"  Late deliveries  : {total_late}")
    if total_late == 0:
        print("  ✅ All deliveries within time windows")
    else:
        print(f"  ⚠  {total_late} deliveries missed their time window")
    print("="*70)


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":

    # Step 1: Geocode all addresses → coordinates
    locations = geocode_all_cached(ORDERS, WAREHOUSE)

    # Step 2: Real road distance matrix (one API call)
    distances, durations = get_distance_matrix(locations)
    if distances is None:
        exit(1)

    # Step 3: Assign orders to vehicles
    vehicle_orders = assign_orders_to_vehicles(ORDERS)

    # Step 4: Assign drivers to vehicles
    driver_assignment = assign_drivers(vehicle_orders)

    # Step 5: Build routes for each vehicle
    print("\n" + "="*65)
    print("  STEP 4 — ROUTING (Priority Lock + Clarke-Wright + 2-opt)")
    print("="*65)
    print("  Rule: Priority beats proximity. Always.")

    loc_index      = {loc["id"]: i for i, loc in enumerate(locations)}
    vehicle_routes = []

    for v in VEHICLES:
        orders = vehicle_orders.get(v["id"], [])
        if not orders:
            continue

        driver = driver_assignment.get(v["id"])
        dname  = driver["name"] if driver else "No driver"

        stop_indices = [loc_index[o["id"]] for o in orders if o["id"] in loc_index]

        print(f"\n  {v['label']} ({dname}) — {len(stop_indices)} stops")

        # Build prioritized + optimized route
        final_route, locked_count = build_route(stop_indices, locations, distances)

        print(f"    Locked (priority): {[locations[i]['name'] for i in final_route[:locked_count]]}")
        print(f"    Optimized (normal): {[locations[i]['name'] for i in final_route[locked_count:]]}")
        print(f"    Final order: {[locations[i]['name'] for i in final_route]}")

        # Total distance
        total_km = distances[0][final_route[0]]
        for k in range(len(final_route) - 1):
            total_km += distances[final_route[k]][final_route[k + 1]]
        total_km += distances[final_route[-1]][0]

        # Simulate timing
        sim_stops, total_min = simulate_route(final_route, locations, durations, driver or DRIVERS[-1], distances)

        # Distance from last stop back to warehouse
        return_km = distances[final_route[-1]][0] if final_route else 0

        vehicle_routes.append({
            "vehicle":            v,
            "driver":             driver,
            "stops":              sim_stops,
            "total_km":           total_km,
            "total_duration_min": total_min,
            "return_km":          round(return_km, 2),
        })

    # Step 6: Print final delivery plan
    print_routes(vehicle_routes)
